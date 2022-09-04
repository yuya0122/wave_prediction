import os
import openpyxl
from collections import defaultdict

INPUT_DIR = "input"
OUTPUT_DIR = "output"
INPUT_EXCEL = os.path.join(INPUT_DIR, "table_design.xlsx")
DB = 1
SCHEMA = 2
TABLE = 3
COLUMN = 4
TYPE = 5
PK = 6
NOT_NULL = 7
START_ROW = 2


def generate_ddl():
    
    # EXCEL to dict
    wb = openpyxl.load_workbook(INPUT_EXCEL)
    ws = wb["table_design"]
    max_row = ws.max_row
    recursive_defaultdict = lambda: defaultdict(recursive_defaultdict)
    param_dict = recursive_defaultdict()
    for i in range(START_ROW, max_row+1):
        db = ws.cell(row=i, column=DB).value
        schema = ws.cell(row=i, column=SCHEMA).value
        table = ws.cell(row=i, column=TABLE).value
        column = ws.cell(row=i, column=COLUMN).value
        type = ws.cell(row=i, column=TYPE).value
        pk = ws.cell(row=i, column=PK).value
        if pk == "Y":
            pk = True
        else:
            pk = False
        not_null = ws.cell(row=i, column=NOT_NULL).value
        if not_null == "Y":
            not_null = True
        else:
            not_null = False
        param_dict[table]["db"] = db
        param_dict[table]["schema"] = schema
        param_dict[table]["columns"][column]["type"] = type
        param_dict[table]["columns"][column]["pk"] = pk
        param_dict[table]["columns"][column]["not_null"] = not_null
        
    # dict to file  
    for table, table_info_dict in param_dict.items():
        db = table_info_dict["db"]
        schema = table_info_dict["schema"]
        columns = table_info_dict["columns"]
        # create tableの行作成
        create_table_row = f"CREATE TABLE {schema}.{table}(\n"
        # column定義の行作成
        column_def_rows = []
        pk_columns = []
        for column_name, type_pk_null_dict in columns.items():
            type = type_pk_null_dict["type"]
            pk = type_pk_null_dict["pk"]
            not_null = type_pk_null_dict["not_null"]
            
            if not_null:
                column_def_row = f"{column_name} {type} NOT NULL"
            else:
                column_def_row = f"{column_name} {type}"
            column_def_rows.append(column_def_row)
            
            if pk:
                pk_columns.append(column_name)
        column_def_rows = "\n,".join(column_def_rows)

        # PK定義の行作成
        if pk_columns:
            pk_columns = ",".join(pk_columns)
            pk_def_row = f"\n,PRIMARY KEY({pk_columns})\n);"
        else:
            pk_def_row = "\n);"

        # 全ての行の結合
        ddl_text = create_table_row + column_def_rows + pk_def_row
        
        # DDLファイル出力
        file = os.path.join(OUTPUT_DIR, f"{table}.sql")
        with open(file, mode="w") as f:
            f.write(ddl_text)

        
if __name__ == "__main__":
    generate_ddl()
            
        


        

    
    
    